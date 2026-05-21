# -*- coding: utf-8 -*-
"""
货盘沉淀脚本
- 输入：D:\1.work\5-20\货盘\*.xlsx
- 输出：
  - feishu_import\08_pallet_flat.csv           扁平化货品表
  - feishu_import\09_pallet_modules_discovered.csv  从升级思路中挖出的模块候选
- 编码：UTF-8 with BOM（Excel 双击不乱码）
"""
import openpyxl, os, csv, re, sys, datetime

SRC_DIR = r"D:\1.work\5-20\货盘"
OUT_DIR = r"D:\1.work\5-20\feishu_import"

FIXED_HEADERS = {"一级品类", "二级品类", "细分品类（按版型）", "后续升级思路", "参考", "备注"}
KNOWN_BRANDS  = {"TMT", "SERUNA", "JAFFICK", "ANTA"}  # 出现在列头里的认作品牌列


def find_header_row(ws):
    """在前 5 行内找含"一级品类"的那一行作为表头。返回 (row_idx_1based, headers_dict {col_idx_1based: header_text})"""
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=5), start=1):
        if any(v == "一级品类" for v in row if v is not None):
            headers = {j + 1: (str(v).strip() if v is not None else None) for j, v in enumerate(row)}
            return i, headers
    return None, None


def classify_columns(headers):
    """根据表头给出列角色：
       l1_col, l2_col, sub_col, age_col(代际), field_col(字段名), brand_cols {col: brand}, upgrade_col, ref_pairs [(参考col,备注col)..]
    """
    l1 = l2 = sub = up = None
    brand_cols = {}
    ref_pairs = []
    last_ref = None
    for c, h in sorted(headers.items()):
        if h == "一级品类":       l1 = c
        elif h == "二级品类":     l2 = c
        elif h == "细分品类（按版型）": sub = c
        elif h == "后续升级思路": up = c
        elif h in KNOWN_BRANDS:   brand_cols[c] = h
        elif h == "参考":         last_ref = c
        elif h == "备注" and last_ref is not None:
            ref_pairs.append((last_ref, c)); last_ref = None
    # 代际列、字段名列推断：在 sub_col 和最早的 brand_col 之间的两列分别是 代际、字段名
    first_brand = min(brand_cols.keys()) if brand_cols else None
    age_col = field_col = None
    if sub is not None and first_brand is not None and first_brand - sub >= 2:
        age_col = sub + 1
        field_col = sub + 2
    return l1, l2, sub, age_col, field_col, brand_cols, up, ref_pairs


def is_image_formula(v):
    if v is None: return False
    s = str(v)
    return s.startswith("=DISPIMG") or s.startswith("=_xlfn.DISPIMG") or "DISPIMG" in s


def clean_cell(v):
    if v is None: return ""
    if is_image_formula(v): return ""
    return str(v).replace("\r", " ").replace("\n", " / ").strip()


def pick_latest_month_sheet(wb):
    """选择最新月份的工作表。优先匹配 26年04月 / 2026.04 等格式。"""
    pat = re.compile(r"(\d{2,4})[年.\-/](\d{1,2})")
    candidates = []
    for sn in wb.sheetnames:
        m = pat.search(sn)
        if m:
            y, mo = m.groups()
            y = int(y); mo = int(mo)
            if y < 100: y += 2000
            candidates.append((y, mo, sn))
    if not candidates: return None
    candidates.sort(reverse=True)
    return candidates[0][2]


# ---------- 模块发现关键词库 ----------
MODULE_KEYWORDS = {
    "材料模块": [
        "硅胶", "EVA", "PU", "TPE", "海绵", "记忆棉", "竹纤维", "竹炭", "羊毛", "羊绒",
        "蚕丝", "莫代尔", "莱卡", "氨纶", "锦纶", "尼龙", "聚酯", "涤纶", "网布", "3D网",
        "泡棉", "凝胶", "石墨烯", "亲肤", "抗菌", "银离子", "远红外", "蜂巢", "羊角绒"
    ],
    "结构模块": [
        "3D编织", "编织", "热压", "绑带", "魔术贴", "弹簧", "支撑条", "钢条", "铝条",
        "护板", "分区", "立体剪裁", "立裁", "工字", "X型", "Y型", "螺旋", "包覆", "绑绳",
        "压胶", "缝合线", "压胶条", "内折", "一片式", "无缝", "拼接", "罩杯", "插片",
        "硅胶垫", "支撑结构", "山型"
    ],
    "功能模块": [
        "防滑", "防臭", "透气", "吸汗", "加压", "恒温", "磁疗", "反光", "发热",
        "降温", "塑形", "收腹", "防勒", "压力均衡", "助眠", "安神", "减压", "助眠枕",
        "颈椎支撑", "防滑点阵"
    ],
    "外观模块": [
        "撞色", "拼色", "印花", "刺绣", "logo绣", "条纹", "渐变"
    ],
    "包装模块": [
        "礼盒", "袋装", "盒装", "彩盒", "PVC包装"
    ],
    "版型模块": [
        "前绑", "后绑", "无绑带", "双层", "单层", "长款", "短款",
        "高腰", "中腰", "低腰", "V领", "圆领", "环膝", "环踝", "环腕"
    ],
}


def discover_modules(text, category, brand):
    if not text: return []
    found = []
    for dim, kws in MODULE_KEYWORDS.items():
        for kw in kws:
            if kw in text:
                found.append((dim, kw))
    return found


def parse_workbook(path):
    fn = os.path.basename(path)
    cat_name = fn.replace(".xlsx", "").replace("货盘", "").strip("()0123456789").strip() or "未知"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sn = pick_latest_month_sheet(wb)
    if not sn:
        print(f"  [skip] {fn}: 找不到月份sheet"); wb.close(); return [], [], None
    ws = wb[sn]
    h_row, headers = find_header_row(ws)
    if not h_row:
        print(f"  [skip] {fn}/{sn}: 找不到表头"); wb.close(); return [], [], sn
    l1, l2, sub, age_col, field_col, brand_cols, up_col, ref_pairs = classify_columns(headers)
    if not (sub and field_col and brand_cols):
        print(f"  [skip] {fn}/{sn}: 表头缺关键列 sub={sub} field={field_col} brands={brand_cols}")
        wb.close(); return [], [], sn

    rows = list(ws.iter_rows(values_only=True, min_row=h_row + 1))
    wb.close()

    state = {"l1": cat_name, "l2": "", "sub": "", "age": ""}
    blocks = {}        # key=(l2, sub, age) → {brand: {货号, 销量, 升级情况}}
    upgrade_global = {}  # key=(l2, sub) → 后续升级思路（按二级品类）

    def get(row, c):
        if c is None or c - 1 >= len(row): return None
        return row[c - 1]

    for ridx, row in enumerate(rows, start=h_row + 1):
        v_l1   = clean_cell(get(row, l1)) if l1 else ""
        v_l2   = clean_cell(get(row, l2)) if l2 else ""
        v_sub  = clean_cell(get(row, sub))
        v_age  = clean_cell(get(row, age_col)) if age_col else ""
        v_fld  = clean_cell(get(row, field_col))
        v_up   = clean_cell(get(row, up_col)) if up_col else ""

        if v_l1: state["l1"] = v_l1
        if v_l2: state["l2"] = v_l2
        if v_sub:
            state["sub"] = v_sub
            state["age"] = ""    # 切换细分品类，代际清空等下一行
        if v_age: state["age"] = v_age

        if not state["sub"]: continue
        # 后续升级思路一般写在细分品类首行，按二级品类汇总
        if v_up:
            key2 = (state["l2"], state["sub"])
            upgrade_global.setdefault(key2, [])
            if v_up not in upgrade_global[key2]:
                upgrade_global[key2].append(v_up)

        if not v_fld: continue  # 没有字段名行（如分隔空行）跳过
        if not state["age"]: state["age"] = "未标注"
        key = (state["l2"], state["sub"], state["age"])
        if key not in blocks:
            blocks[key] = {b: {"货号": "", "销量": "", "升级情况": "", "竞品参考": []} for b in brand_cols.values()}
            # 给所有品牌占位

        # 各品牌列的值
        for col, brand in brand_cols.items():
            val = clean_cell(get(row, col))
            if not val: continue
            if v_fld == "图片":   pass  # 丢图片
            elif v_fld == "货号":   blocks[key][brand]["货号"] = val
            elif v_fld == "销量":   blocks[key][brand]["销量"] = val
            elif "升级" in v_fld:   blocks[key][brand]["升级情况"] = val
            else:
                # 兜底：把其他字段塞进升级情况
                blocks[key][brand]["升级情况"] = (blocks[key][brand]["升级情况"] + " | " + f"{v_fld}:{val}").strip(" |")

        # 竞品参考列：抓 URL
        for ref_c, note_c in ref_pairs:
            note = clean_cell(get(row, note_c))
            if note and ("http" in note or "tmall" in note or "jd.com" in note or "taobao" in note):
                # 把链接塞到当前 key 的所有自家品牌作为竞品参考
                for brand in brand_cols.values():
                    if note not in blocks[key][brand]["竞品参考"]:
                        blocks[key][brand]["竞品参考"].append(note)

    # 输出扁平表
    flat_rows = []
    for (l2, subv, age), per_brand in blocks.items():
        for brand, d in per_brand.items():
            if not (d["货号"] or d["销量"] or d["升级情况"]): continue  # 全空跳过
            # 判定状态
            sales = d["销量"]
            if "打样" in sales or "概念" in sales or "暂无" in sales:
                status = "打样/概念"
            elif re.search(r"\d", sales):
                status = "已量产"
            else:
                status = "未知"
            up_text = " / ".join(upgrade_global.get((l2, subv), []))
            flat_rows.append({
                "品类": cat_name, "二级品类": l2, "版型(细分品类)": subv, "代际": age,
                "品牌": brand, "货号": d["货号"], "销量": d["销量"], "状态": status,
                "本代升级情况": d["升级情况"], "细分品类后续升级思路": up_text,
                "竞品参考链接": " | ".join(d["竞品参考"]),
                "数据来源": fn, "数据来源sheet": sn,
            })

    # 输出模块发现表
    module_rows = []
    seen = set()
    tmp_seq = 0
    def add_mod(dim, kw, brand, src_sku, evidence, l2v, subv, age):
        nonlocal tmp_seq
        key = (dim, kw, brand, cat_name)
        if key in seen: return
        seen.add(key); tmp_seq += 1
        module_rows.append({
            "临时编号": f"TMP-PAL-{cat_name[:2]}-{tmp_seq:03d}",
            "模块名": f"{kw}（{cat_name}-{subv}）" if subv else f"{kw}（{cat_name}）",
            "模块维度": dim.replace("模块", ""),
            "归属品牌": brand,
            "一句话说明": f"在{cat_name}{subv}{age}中出现的{dim[:-2]}关键词，原文：{evidence[:80]}",
            "来源SKU": src_sku or f"{cat_name}-{subv}-{age}",
            "状态": "已量产" if src_sku else "概念",
            "关键参数": "",
            "适用品类": cat_name,
            "不适用场景": "",
            "负责人": "党英轩",
            "建议归入维度": "",
            "收集来源": f"货盘沉淀-{fn}",
            "收集时间": datetime.date.today().isoformat(),
            "确认状态": "待负责人确认",
        })

    for (l2v, subv, age), per_brand in blocks.items():
        for brand, d in per_brand.items():
            evidence_pool = [d["升级情况"], subv, age]
            for ev in evidence_pool:
                for dim, kw in discover_modules(ev or "", cat_name, brand):
                    add_mod(dim, kw, brand, d["货号"], ev or "", l2v, subv, age)
        # 二级品类后续升级思路也扫
        up_text = " / ".join(upgrade_global.get((l2v, subv), []))
        if up_text:
            for dim, kw in discover_modules(up_text, cat_name, ""):
                add_mod(dim, kw, "通用", "", up_text, l2v, subv, "")

    return flat_rows, module_rows, sn


def write_csv_utf8_bom(path, rows, fieldnames):
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)


def main():
    sys.stdout.reconfigure(encoding='utf-8')
    all_flat, all_mods = [], []
    report = []
    for fn in sorted(os.listdir(SRC_DIR)):
        if not fn.endswith(".xlsx"): continue
        p = os.path.join(SRC_DIR, fn)
        print(f"\n----- {fn} -----")
        flat, mods, sn = parse_workbook(p)
        print(f"  sheet: {sn}  扁平行={len(flat)}  发现模块={len(mods)}")
        report.append({"文件": fn, "来源sheet": sn or "-", "扁平条数": len(flat), "模块条数": len(mods)})
        all_flat.extend(flat); all_mods.extend(mods)

    fn1 = os.path.join(OUT_DIR, "08_pallet_flat.csv")
    write_csv_utf8_bom(fn1, all_flat, [
        "品类", "二级品类", "版型(细分品类)", "代际", "品牌", "货号", "销量", "状态",
        "本代升级情况", "细分品类后续升级思路", "竞品参考链接", "数据来源", "数据来源sheet"
    ])
    fn2 = os.path.join(OUT_DIR, "09_pallet_modules_discovered.csv")
    write_csv_utf8_bom(fn2, all_mods, [
        "临时编号", "模块名", "模块维度", "归属品牌", "一句话说明", "来源SKU", "状态",
        "关键参数", "适用品类", "不适用场景", "负责人", "建议归入维度", "收集来源",
        "收集时间", "确认状态"
    ])

    print("\n========== 沉淀报告 ==========")
    print(f"  {'文件':<25} {'sheet':<12} {'扁平':>6} {'模块':>6}")
    for r in report:
        print(f"  {r['文件']:<25} {r['来源sheet']:<12} {r['扁平条数']:>6} {r['模块条数']:>6}")
    print(f"\n  TOTAL flat={len(all_flat)}  modules={len(all_mods)}")
    print(f"\n输出：\n  {fn1}\n  {fn2}")


if __name__ == "__main__":
    main()
