# -*- coding: utf-8 -*-
"""
工厂模块库解析 · Step 2 of RawData 重建
─────────────────────────────────────────────
输入：
  RawData/Module File Lib-(Organize files by factory)/博凯模块库.xlsx
  RawData/Module File Lib-(Organize files by factory)/宏博模块库(版型模块补充中)(1).xlsx
输出：
  data/modules.csv               主表（每模块一行）
  data/module_product_link.csv   复用关系（每条复用一行）
  openclaw_output/data_quality_modules.md  质量报告

策略：
- 7 个 sheet × 2 个工厂 = 14 sheets
- R1+R2 二级表头，模块基本字段在 R1，复用记录子字段在 R2
- R3 起：每个模块条目可能跨多行（多条复用记录）
- 「模块编号」非空 = 新模块开始
- 自适应：根据 R1+R2 文本动态找列位置（不写死列号）
"""

import openpyxl
import csv
import os
import sys
import re
import datetime
from collections import Counter

SRC_FILES = [
    ("博凯", r"RawData/Module File Lib-(Organize files by factory)/博凯模块库.xlsx"),
    ("宏博", r"RawData/Module File Lib-(Organize files by factory)/宏博模块库(版型模块补充中)(1).xlsx"),
]

OUT_MODULES = "data/modules.csv"
OUT_LINKS = "data/module_product_link.csv"
OUT_REPORT = "openclaw_output/data_quality_modules.md"

# 列字段的模糊匹配规则（中文换行/全角差异都能匹配上）
def matches(actual, *patterns):
    if not actual: return False
    a = str(actual).replace("\n", "").replace(" ", "").strip()
    return any(p in a for p in patterns)


def is_image_formula(v):
    if v is None: return False
    s = str(v)
    return "DISPIMG" in s


def extract_dispimg_id(v):
    """从 =DISPIMG("ID_XXX",1) 抽 ID_XXX"""
    if not v: return ""
    s = str(v)
    m = re.search(r'DISPIMG\(\s*"([^"]+)"', s)
    return m.group(1) if m else ""


def clean(v):
    if v is None: return ""
    if is_image_formula(v): return ""
    return str(v).replace("\r", " ").replace("\n", " ").strip()


def build_col_map(r1, r2):
    """构建列索引 → 字段角色的映射
    返回 dict: {role: col_index_1based}
    role 取值：序号, 模块编号, 模块类型, 模块名称, 模块尺寸, 生产工厂, 设计时间,
              使用材料, 颜色, 价格, 备注, 正面截图, 背面截图,
              复用序号, 复用产品名, 复用产品编号, 复用位置, 复用产品截图, 复用备注
    """
    col_map = {}

    def set_if(col_idx, role, value, *patterns):
        if matches(value, *patterns) and role not in col_map:
            col_map[role] = col_idx

    # 处理 R1
    for i, v in enumerate(r1 or [], start=1):
        set_if(i, "序号", v, "序号")
        set_if(i, "模块编号", v, "模块编号")
        set_if(i, "模块类型", v, "模块类型")
        set_if(i, "模块名称", v, "模块名称")
        set_if(i, "模块尺寸", v, "模块尺寸")
        set_if(i, "生产工厂", v, "生产工厂")
        set_if(i, "设计时间", v, "设计时间")
        set_if(i, "使用材料", v, "使用材料", "材料")
        set_if(i, "颜色", v, "颜色")
        set_if(i, "价格", v, "价格", "erp成本", "erp")
        set_if(i, "备注", v, "备注")
        set_if(i, "正面截图", v, "正面")
        set_if(i, "背面截图", v, "背面")

    # 处理 R2 (复用记录子字段)
    for i, v in enumerate(r2 or [], start=1):
        set_if(i, "复用序号", v, "复用序号")
        set_if(i, "复用产品名", v, "产品名称")
        set_if(i, "复用产品编号", v, "产品编号")
        set_if(i, "复用位置", v, "复用位置")
        set_if(i, "复用产品截图", v, "产品截图", "穿戴图")
        # R2 的"备注"列单独识别（避免和 R1 的"备注"冲突）
        if matches(v, "备注") and i > col_map.get("复用产品截图", 0):
            col_map["复用备注"] = i

    return col_map


def parse_sheet(factory, sheet_name, ws):
    """解析单个 sheet
    返回：(modules_list, links_list)
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return [], []

    r1, r2 = rows[0], rows[1]
    col_map = build_col_map(r1, r2)

    if "模块编号" not in col_map:
        print(f"  ⚠️ {factory}/{sheet_name}: 找不到「模块编号」列，跳过")
        return [], []

    def get(row, role):
        col = col_map.get(role)
        if col is None or col - 1 >= len(row): return None
        return row[col - 1]

    modules = []
    links = []
    current_module = None
    today = datetime.date.today().isoformat()

    for ridx, row in enumerate(rows[2:], start=3):
        if row is None or all(v is None or v == "" for v in row):
            continue

        mod_id_raw = clean(get(row, "模块编号"))
        is_new_module = bool(mod_id_raw)

        if is_new_module:
            # 新模块开始
            mod_id = mod_id_raw
            current_module = {
                "module_id": mod_id,
                "module_type_sheet": sheet_name,
                "module_type_col": clean(get(row, "模块类型")),
                "module_name": clean(get(row, "模块名称")) or clean(get(row, "使用材料")),
                "size": clean(get(row, "模块尺寸")),
                "factory_src": factory,
                "factory_col": clean(get(row, "生产工厂")),
                "design_time": clean(get(row, "设计时间")),
                "material": clean(get(row, "使用材料")),
                "color": clean(get(row, "颜色")),
                "price": clean(get(row, "价格")),
                "remark": clean(get(row, "备注")),
                "image_front_id": extract_dispimg_id(get(row, "正面截图")),
                "image_back_id": extract_dispimg_id(get(row, "背面截图")),
                "source_sheet": sheet_name,
                "source_file": f"{factory}模块库",
                "row_in_sheet": ridx,
                "ingested_at": today,
            }
            modules.append(current_module)

        # 不管是不是新模块，都尝试解析复用记录（同一模块可能跨多行）
        if current_module is not None:
            reuse_idx = clean(get(row, "复用序号"))
            prod_name = clean(get(row, "复用产品名"))
            prod_code = clean(get(row, "复用产品编号"))
            reuse_pos = clean(get(row, "复用位置"))
            prod_img = extract_dispimg_id(get(row, "复用产品截图"))
            reuse_remark = clean(get(row, "复用备注"))

            # 只要有任何一个复用字段非空，就记一条复用关系
            if any([reuse_idx, prod_name, prod_code, reuse_pos, prod_img]):
                links.append({
                    "module_id": current_module["module_id"],
                    "reuse_idx": reuse_idx,
                    "product_name": prod_name,
                    "product_code": prod_code,
                    "reuse_position": reuse_pos,
                    "product_image_id": prod_img,
                    "remark": reuse_remark,
                    "factory_src": factory,
                    "module_type_sheet": sheet_name,
                })

    return modules, links


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    os.makedirs("data", exist_ok=True)
    os.makedirs("openclaw_output", exist_ok=True)

    all_modules = []
    all_links = []
    sheet_stats = []

    for factory, path in SRC_FILES:
        print(f"\n===== {factory} ({os.path.basename(path)}) =====")
        if not os.path.exists(path):
            print(f"  ❌ 文件不存在")
            continue

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            if sn.startswith("Wps"): continue
            ws = wb[sn]
            mods, links = parse_sheet(factory, sn, ws)
            print(f"  {sn:14s}  rows={ws.max_row:5d}  → 模块 {len(mods):4d}  复用 {len(links):4d}")
            sheet_stats.append({"工厂": factory, "Sheet": sn, "原始行数": ws.max_row,
                                 "模块数": len(mods), "复用记录数": len(links)})
            all_modules.extend(mods)
            all_links.extend(links)
        wb.close()

    # 写出 modules.csv
    mod_fields = ["module_id", "module_type_sheet", "module_type_col", "module_name",
                  "size", "factory_src", "factory_col", "design_time", "material",
                  "color", "price", "remark", "image_front_id", "image_back_id",
                  "source_sheet", "source_file", "row_in_sheet", "ingested_at"]
    with open(OUT_MODULES, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=mod_fields)
        w.writeheader()
        w.writerows(all_modules)

    # 写出 module_product_link.csv
    link_fields = ["module_id", "reuse_idx", "product_name", "product_code",
                   "reuse_position", "product_image_id", "remark",
                   "factory_src", "module_type_sheet"]
    with open(OUT_LINKS, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=link_fields)
        w.writeheader()
        w.writerows(all_links)

    # ===== 质量统计 =====
    mod_by_type = Counter(m["module_type_sheet"] for m in all_modules)
    mod_by_factory = Counter(m["factory_src"] for m in all_modules)
    links_by_factory = Counter(l["factory_src"] for l in all_links)

    # 一个模块平均被复用几次
    reuse_count_per_module = Counter(l["module_id"] for l in all_links)
    high_reuse = sorted(reuse_count_per_module.items(), key=lambda x: -x[1])[:15]

    # 产品编号填充情况
    with_product_code = sum(1 for l in all_links if l["product_code"])
    with_product_name = sum(1 for l in all_links if l["product_name"])

    # ===== 写报告 =====
    today = datetime.date.today().isoformat()
    lines = [
        f"# 工厂模块库解析 · 数据质量报告",
        f"",
        f"> 生成时间：{today}",
        f"> 输入：博凯+宏博 2 个 xlsx × 7 sheets",
        f"> 输出：",
        f"> - `{OUT_MODULES}`",
        f"> - `{OUT_LINKS}`",
        f"",
        f"## 📊 总览",
        f"",
        f"| 指标 | 值 |",
        f"|---|---:|",
        f"| 模块总数 | **{len(all_modules)}** |",
        f"| 模块复用记录总数 | **{len(all_links)}** |",
        f"| 平均每模块被复用次数 | {len(all_links)/len(all_modules):.2f} |" if all_modules else "| 平均每模块复用 | — |",
        f"| 复用记录含产品编号 | {with_product_code}（{with_product_code/len(all_links)*100:.1f}%）|" if all_links else "",
        f"| 复用记录含产品名称 | {with_product_name}（{with_product_name/len(all_links)*100:.1f}%）|" if all_links else "",
        f"",
        f"## 🏭 按工厂分布",
        f"",
        f"| 工厂 | 模块数 | 复用记录数 |",
        f"|---|---:|---:|",
    ]
    for fct in ["博凯", "宏博"]:
        lines.append(f"| {fct} | {mod_by_factory.get(fct, 0)} | {links_by_factory.get(fct, 0)} |")

    lines.extend([
        f"",
        f"## 📦 按模块类型分布",
        f"",
        f"| 类型 | 模块数 |",
        f"|---|---:|",
    ])
    for mtype, cnt in mod_by_type.most_common():
        lines.append(f"| {mtype} | {cnt} |")

    lines.extend([
        f"",
        f"## 📑 各 Sheet 解析明细",
        f"",
        f"| 工厂 | Sheet | 原始行数 | 解析出模块 | 解析出复用 |",
        f"|---|---|---:|---:|---:|",
    ])
    for s in sheet_stats:
        lines.append(f"| {s['工厂']} | {s['Sheet']} | {s['原始行数']} | {s['模块数']} | {s['复用记录数']} |")

    lines.extend([
        f"",
        f"## 🏆 被复用最多的 15 个模块（金贵资产）",
        f"",
        f"| 排名 | 模块编号 | 被复用次数 |",
        f"|---:|---|---:|",
    ])
    for i, (mid, cnt) in enumerate(high_reuse, start=1):
        lines.append(f"| {i} | `{mid}` | {cnt} |")

    lines.extend([
        f"",
        f"## ⚠️ 已知问题与下一步",
        f"",
        f"1. **图片字段是 DISPIMG ID（飞书图床）**——`image_front_id` / `image_back_id` / `product_image_id`",
        f"   都只保存了 ID，迁出原表后无法预览。v1 先留 ID 占位，v2 考虑批量导出图片到 OSS。",
        f"2. **版型/魔术贴/外观 sheet 没有「产品编号」字段**——这些类别的复用记录里",
        f"   `product_code` 字段为空，只能靠 `product_name` 做模糊 JOIN 到 `product_assets.csv`。",
        f"3. **DISPIMG 图片整体丢失约 {sum(1 for m in all_modules if not m['image_front_id'])} 个模块的正面截图**",
        f"   （这些行原本可能就没图，或图片是嵌入式不是公式式）。",
        f"",
        f"## 🔄 下一步",
        f"",
        f"- Step 3：已经完成（modules.csv + module_product_link.csv 同步产出）✅",
        f"- Step 4：重写 sediment_pallet.py，只产 `data/competitor_intel.csv`",
        f"- Step 5：交叉验证 4 张表 JOIN 通不通（特别是 module_product_link.product_code ↔ product_assets.货品编号）",
    ])

    with open(OUT_REPORT, "w", encoding="utf-8-sig") as f:
        f.write("\n".join(lines))

    print()
    print(f"✅ Step 2+3 完成")
    print(f"  - {OUT_MODULES}  ({len(all_modules)} 模块)")
    print(f"  - {OUT_LINKS}    ({len(all_links)} 复用记录)")
    print(f"  - {OUT_REPORT}")
    print()
    print(f"📊 速览：")
    print(f"  博凯={mod_by_factory.get('博凯', 0)} 模块 / 宏博={mod_by_factory.get('宏博', 0)} 模块")
    print(f"  平均每模块被复用 {len(all_links)/len(all_modules):.2f} 次" if all_modules else "")
    if high_reuse:
        print(f"  最金贵模块：{high_reuse[0][0]}（被复用 {high_reuse[0][1]} 次）")


if __name__ == "__main__":
    main()
