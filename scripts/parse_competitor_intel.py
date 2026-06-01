# -*- coding: utf-8 -*-
"""
货盘竞品情报提取 · Step 4 of RawData 重建
─────────────────────────────────────────────
输入：RawData/货盘/*.xlsx  (4 个：护腕/护膝/护踝/睡眠)
输出：
  data/competitor_intel.csv         竞品情报表
  openclaw_output/data_quality_competitor.md  质量报告

策略：
- 旧 sediment_pallet.py 同时产「自家 SKU 扁平表」+「模块候选反推」+「竞品参考」
  但前两者已被销量表 + 工厂模块库替代，不再生成
- 本脚本只保留「参考/备注」列里的竞品 URL + 销量描述 + 团队升级思路
- 每行 = 一个细分品类×代际块下的一个竞品参考槽
"""

import openpyxl
import os
import csv
import re
import sys
import datetime
from collections import Counter

SRC_DIR = r"RawData/货盘"
OUT_CSV = "data/competitor_intel.csv"
OUT_REPORT = "openclaw_output/data_quality_competitor.md"

KNOWN_BRANDS = {"TMT", "SERUNA", "JAFFICK", "ANTA"}


def is_image_formula(v):
    if v is None: return False
    return "DISPIMG" in str(v)


def clean(v):
    if v is None: return ""
    if is_image_formula(v): return ""
    return str(v).replace("\r", " ").replace("\n", " / ").strip()


def is_url(s):
    if not s: return False
    return any(k in s for k in ("http://", "https://", "tmall.com", "taobao.com", "jd.com", "douyin.com", "xiaohongshu", "1688"))


def find_header_row(ws):
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=5), start=1):
        if any(v == "一级品类" for v in row if v is not None):
            return i, {j + 1: (str(v).strip() if v is not None else None) for j, v in enumerate(row)}
    return None, None


def classify_columns(headers):
    l1 = l2 = sub = up = None
    brand_cols = {}
    ref_pairs = []
    last_ref = None
    for c, h in sorted(headers.items()):
        if h == "一级品类":       l1 = c
        elif h == "二级品类":     l2 = c
        elif h == "细分品类（按版型）": sub = c
        elif h == "后续升级思路": up = c
        elif h in KNOWN_BRANDS:   brand_cols[c] = h
        elif h == "参考":         last_ref = c
        elif h == "备注" and last_ref is not None:
            ref_pairs.append((last_ref, c)); last_ref = None
    first_brand = min(brand_cols.keys()) if brand_cols else None
    age_col = field_col = None
    if sub is not None and first_brand is not None and first_brand - sub >= 2:
        age_col = sub + 1
        field_col = sub + 2
    return l1, l2, sub, age_col, field_col, brand_cols, up, ref_pairs


def pick_latest_month_sheet(wb):
    pat = re.compile(r"(\d{2,4})[年.\-/](\d{1,2})")
    cands = []
    for sn in wb.sheetnames:
        m = pat.search(sn)
        if m:
            y, mo = int(m.group(1)), int(m.group(2))
            if y < 100: y += 2000
            cands.append((y, mo, sn))
    cands.sort(reverse=True)
    return cands[0][2] if cands else None


def parse_workbook(path):
    fn = os.path.basename(path)
    cat = fn.replace(".xlsx", "").replace("货盘", "").strip("()0123456789").strip() or "未知"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sn = pick_latest_month_sheet(wb)
    if not sn:
        wb.close(); return [], cat, None

    ws = wb[sn]
    h_row, headers = find_header_row(ws)
    if not h_row:
        wb.close(); return [], cat, sn

    l1, l2, sub, age_col, _field_col, brand_cols, up_col, ref_pairs = classify_columns(headers)
    if not (sub and ref_pairs):
        wb.close(); return [], cat, sn

    rows = list(ws.iter_rows(values_only=True, min_row=h_row + 1))
    wb.close()

    def get(row, c):
        if c is None or c - 1 >= len(row): return None
        return row[c - 1]

    # 状态追踪：当前所在的 l1/l2/sub/age
    state = {"l1": cat, "l2": "", "sub": "", "age": ""}
    # 当前 sub 块下的所有竞品 URL（多行累计）
    blocks = {}  # key=(l2, sub, age) → {"urls": [], "notes": [], "upgrades": [], "our_brand_codes": {brand: codes}}

    for row in rows:
        if row is None: continue
        v_l1   = clean(get(row, l1)) if l1 else ""
        v_l2   = clean(get(row, l2)) if l2 else ""
        v_sub  = clean(get(row, sub))
        v_age  = clean(get(row, age_col)) if age_col else ""
        v_up   = clean(get(row, up_col)) if up_col else ""

        if v_l1: state["l1"] = v_l1
        if v_l2: state["l2"] = v_l2
        if v_sub:
            state["sub"] = v_sub
            state["age"] = ""
        if v_age: state["age"] = v_age

        if not state["sub"]: continue
        if not state["age"]: state["age"] = "未标注"

        key = (state["l2"], state["sub"], state["age"])
        if key not in blocks:
            blocks[key] = {"urls": [], "notes": [], "upgrades": [], "our_brand_codes": {}}

        # 累计升级思路（同一细分品类可能多行）
        if v_up and v_up not in blocks[key]["upgrades"]:
            blocks[key]["upgrades"].append(v_up)

        # 扫描每对 参考/备注 列
        for slot_idx, (ref_c, note_c) in enumerate(ref_pairs, start=1):
            note = clean(get(row, note_c))
            if is_url(note):
                if note not in blocks[key]["urls"]:
                    blocks[key]["urls"].append((slot_idx, note))
            elif note and not is_url(note):
                # 备注列也可能是销量描述如"3000＋ / 单链接双商品"
                if note not in blocks[key]["notes"]:
                    blocks[key]["notes"].append((slot_idx, note))

    # 转成扁平输出
    rows_out = []
    for (l2v, subv, age), d in blocks.items():
        upgrade_text = " / ".join(d["upgrades"])
        # 每个 URL 输出一行
        if d["urls"]:
            for slot_idx, url in d["urls"]:
                # 找同 slot 的销量描述
                matching_note = next((n for s, n in d["notes"] if s == slot_idx), "")
                rows_out.append({
                    "pallet_file": fn,
                    "month_sheet": sn,
                    "l1_category": cat,
                    "l2_category": l2v,
                    "sub_category": subv,
                    "generation": age,
                    "ref_slot": slot_idx,
                    "competitor_url": url,
                    "competitor_sales_note": matching_note,
                    "team_upgrade_idea": upgrade_text,
                    "source": "货盘参考列",
                })
        elif upgrade_text:
            # 没有竞品 URL 但有升级思路 → 也保留一行
            rows_out.append({
                "pallet_file": fn,
                "month_sheet": sn,
                "l1_category": cat,
                "l2_category": l2v,
                "sub_category": subv,
                "generation": age,
                "ref_slot": "",
                "competitor_url": "",
                "competitor_sales_note": "",
                "team_upgrade_idea": upgrade_text,
                "source": "货盘升级思路列",
            })

    return rows_out, cat, sn


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    os.makedirs("data", exist_ok=True)
    os.makedirs("openclaw_output", exist_ok=True)

    all_rows = []
    per_file = []
    for fn in sorted(os.listdir(SRC_DIR)):
        if not fn.endswith(".xlsx"): continue
        path = os.path.join(SRC_DIR, fn)
        print(f"\n----- {fn} -----")
        rows, cat, sn = parse_workbook(path)
        urls_cnt = sum(1 for r in rows if r["competitor_url"])
        idea_cnt = sum(1 for r in rows if r["source"] == "货盘升级思路列")
        print(f"  sheet: {sn}  竞品 URL={urls_cnt}  独立升级思路={idea_cnt}")
        per_file.append({"file": fn, "sheet": sn, "urls": urls_cnt, "ideas": idea_cnt, "total": len(rows)})
        all_rows.extend(rows)

    # 写 CSV
    fields = ["pallet_file", "month_sheet", "l1_category", "l2_category", "sub_category",
              "generation", "ref_slot", "competitor_url", "competitor_sales_note",
              "team_upgrade_idea", "source"]
    with open(OUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(all_rows)

    # 质量报告
    by_cat = Counter(r["l1_category"] for r in all_rows)
    url_rows = [r for r in all_rows if r["competitor_url"]]
    with_sales_note = sum(1 for r in url_rows if r["competitor_sales_note"])
    with_upgrade = sum(1 for r in url_rows if r["team_upgrade_idea"])

    today = datetime.date.today().isoformat()
    lines = [
        f"# 货盘竞品情报 · 数据质量报告",
        f"",
        f"> 生成时间：{today}",
        f"> 输入：`RawData/货盘/*.xlsx` (4 个货盘最新月份 sheet)",
        f"> 输出：`{OUT_CSV}`",
        f"",
        f"## 📊 总览",
        f"",
        f"| 指标 | 值 |",
        f"|---|---:|",
        f"| 总记录数 | **{len(all_rows)}** |",
        f"| 含竞品 URL 行数 | {len(url_rows)} |",
        f"| URL 行含销量描述 | {with_sales_note}（{with_sales_note/max(1,len(url_rows))*100:.1f}%）|",
        f"| URL 行含团队升级思路 | {with_upgrade}（{with_upgrade/max(1,len(url_rows))*100:.1f}%）|",
        f"| 仅有升级思路无 URL | {len(all_rows) - len(url_rows)} |",
        f"",
        f"## 📦 按品类分布",
        f"",
        f"| 品类 | 记录数 | 含 URL |",
        f"|---|---:|---:|",
    ]
    for cat, cnt in by_cat.most_common():
        url_in_cat = sum(1 for r in url_rows if r["l1_category"] == cat)
        lines.append(f"| {cat} | {cnt} | {url_in_cat} |")

    lines.extend([
        f"",
        f"## 📑 按货盘文件",
        f"",
        f"| 文件 | sheet | 竞品 URL | 独立升级思路 |",
        f"|---|---|---:|---:|",
    ])
    for f in per_file:
        lines.append(f"| {f['file']} | {f['sheet']} | {f['urls']} | {f['ideas']} |")

    lines.extend([
        f"",
        f"## 🔍 样本（前 5 条竞品 URL）",
        f"",
    ])
    for r in url_rows[:5]:
        lines.append(f"- **[{r['l1_category']}/{r['sub_category']}/{r['generation']}]** ")
        lines.append(f"  - URL: {r['competitor_url'][:90]}...")
        if r['competitor_sales_note']:
            lines.append(f"  - 销量描述: {r['competitor_sales_note']}")
        if r['team_upgrade_idea']:
            lines.append(f"  - 升级思路: {r['team_upgrade_idea'][:80]}")

    lines.extend([
        f"",
        f"## ⚠️ 已知问题",
        f"",
        f"1. URL 去重：同一个 URL 可能在不同行重复出现（货盘是矩阵布局），已做 set 去重。",
        f"2. 部分备注列内容不是 URL 而是销量描述（如「3000＋ / 单链接双商品」）——已分别归到 `competitor_sales_note`。",
        f"3. 团队升级思路按细分品类合并（多代际共享），可能丢失「哪一代升级」的精度。",
        f"",
        f"## 🔄 下一步",
        f"",
        f"- Step 5：交叉验证 4 张表 JOIN 通不通：",
        f"  - module_product_link.product_code ↔ product_assets.货品编号",
        f"  - competitor_intel.sub_category ↔ product_assets.分类（语义层映射）",
    ])

    with open(OUT_REPORT, "w", encoding="utf-8-sig") as f:
        f.write("\n".join(lines))

    print()
    print(f"✅ Step 4 完成")
    print(f"  - {OUT_CSV}  ({len(all_rows)} 行 / 含 URL {len(url_rows)})")
    print(f"  - {OUT_REPORT}")


if __name__ == "__main__":
    main()
