# -*- coding: utf-8 -*-
"""
销量表解析脚本 · Step 1 of RawData 重建
─────────────────────────────────────────────
输入：RawData/Product Monthly Sales File/4月销量汇总表.xlsx
输出：data/product_assets.csv  (UTF-8 BOM, 给 Excel 双击不乱码)
       openclaw_output/data_quality_sales.md  (质量报告)

策略：
- 销量表是 SKU 规格级真实销售数据（4621 行 × 40 列）
- 列名直接保留中文，方便业务侧对照原表
- 增加 1 列 sku_full_id = 货品编号 + 规格码（建主键）
- 增加 1 列 is_zero_sales（标记 0 销量 SKU，未起量候选）
"""

import openpyxl
import csv
import os
import sys
import datetime
from collections import Counter

# ===== 路径 =====
SRC = r"RawData/Product Monthly Sales File/4月销量汇总表.xlsx"
OUT_CSV = r"data/product_assets.csv"
OUT_REPORT = r"openclaw_output/data_quality_sales.md"

# 销量表 40 列的中文表头（按实际顺序）
ORIG_HEADERS = [
    "图片",          # C1
    "商家编码",      # C2  ⭐ 主键候选
    "店铺",          # C3
    "品牌",          # C4
    "分类",          # C5  → category
    "货品编号",      # C6  ⭐ 产品级编号
    "货品名称",      # C7
    "货品简称",      # C8
    "规格码",        # C9
    "规格名称",      # C10
    "均价",          # C11
    "零售价",        # C12
    "批发价",        # C13
    "会员价",        # C14
    "市场价",        # C15
    "最低价",        # C16
    "ERP价格",       # C17
    "自定义价格2",   # C18
    "折扣率",        # C19
    "发货总量",      # C20
    "发货赠品总量",  # C21
    "退货总量",      # C22
    "批次",          # C23
    "批次发货量",    # C24
    "实际销售量",    # C25  ⭐ 核心指标
    "邮资收入",      # C26
    "发货总金额",    # C27
    "未知成本销售总额",  # C28
    "佣金成本",      # C29
    "货品总成本",    # C30
    "货品总利润",    # C31
    "退货总金额",    # C32
    "退货总成本",    # C33
    "发货总金额-发货后实际退款金额",  # C34
    "实际货品总成本",  # C35
    "实际货品总利润",  # C36
    "毛利率%",       # C37  ⭐ 健康度指标
    "发货后实际退款金额",  # C38
    "发货总金额-退货总金额",  # C39
    "单品支付总额",  # C40
]


def to_num(v):
    """尽量转数字；失败返回原值"""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if s == "" or s == "-":
        return None
    try:
        if "." in s:
            return float(s)
        return int(s)
    except ValueError:
        return s


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    print(f"读取：{SRC}")
    if not os.path.exists(SRC):
        print(f"❌ 文件不存在：{SRC}")
        sys.exit(1)

    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    print(f"  Sheet: Sheet1  max_row={ws.max_row}  max_col={ws.max_column}")

    # 读所有行
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter)
    header_list = [str(v).strip() if v else "" for v in header_row]

    # 校验表头
    print(f"  实际表头列数：{len(header_list)}")
    mismatched = []
    for i, expected in enumerate(ORIG_HEADERS):
        if i >= len(header_list):
            mismatched.append((i+1, expected, "MISSING"))
        elif header_list[i] != expected:
            mismatched.append((i+1, expected, header_list[i]))
    if mismatched:
        print(f"⚠️ 表头有 {len(mismatched)} 列不匹配（不致命，继续）：")
        for col, exp, got in mismatched[:5]:
            print(f"    C{col}  期望={exp!r}  实际={got!r}")

    # 写入
    os.makedirs(os.path.dirname(OUT_CSV), exist_ok=True)
    os.makedirs(os.path.dirname(OUT_REPORT), exist_ok=True)

    # 输出 fieldnames：原中文表头 + 2 个增强字段
    out_fields = list(ORIG_HEADERS) + ["sku_full_id", "is_zero_sales"]

    # 统计容器
    stat_total = 0
    stat_by_brand = Counter()
    stat_by_category = Counter()
    stat_by_shop = Counter()
    stat_zero_sales = 0
    stat_no_brand = 0
    stat_no_category = 0
    sales_values = []      # 实际销售量分布
    margin_values = []     # 毛利率分布

    with open(OUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=out_fields)
        w.writeheader()

        for row in rows_iter:
            if row is None: continue
            if all(v is None or v == "" for v in row):
                continue
            stat_total += 1

            rec = {}
            for i, h in enumerate(ORIG_HEADERS):
                if i < len(row):
                    rec[h] = row[i] if row[i] is not None else ""
                else:
                    rec[h] = ""

            # 增强字段
            product_code = str(rec.get("货品编号", "")).strip()
            spec_code = str(rec.get("规格码", "")).strip()
            sku_full_id = f"{product_code}-{spec_code}" if spec_code else product_code
            rec["sku_full_id"] = sku_full_id

            sales = to_num(rec.get("实际销售量"))
            try:
                is_zero = sales is None or float(sales) == 0
            except (TypeError, ValueError):
                is_zero = True
            rec["is_zero_sales"] = "是" if is_zero else "否"

            # 统计
            brand = str(rec.get("品牌", "")).strip()
            category = str(rec.get("分类", "")).strip()
            shop = str(rec.get("店铺", "")).strip()
            if brand: stat_by_brand[brand] += 1
            else: stat_no_brand += 1
            if category: stat_by_category[category] += 1
            else: stat_no_category += 1
            if shop: stat_by_shop[shop] += 1
            if is_zero: stat_zero_sales += 1
            if sales is not None:
                try:
                    sales_values.append(float(sales))
                except (TypeError, ValueError): pass
            margin = to_num(rec.get("毛利率%"))
            if margin is not None:
                try:
                    margin_values.append(float(margin))
                except (TypeError, ValueError): pass

            w.writerow(rec)

    wb.close()

    # ===== 写质量报告 =====
    def pct(num, total):
        return f"{num/total*100:.1f}%" if total else "—"

    def stats(values):
        if not values: return {"count": 0}
        s = sorted(values)
        n = len(s)
        return {
            "count": n,
            "min": s[0],
            "p25": s[n//4] if n >= 4 else s[0],
            "median": s[n//2],
            "p75": s[3*n//4] if n >= 4 else s[-1],
            "max": s[-1],
            "mean": sum(s)/n,
        }

    sales_stats = stats(sales_values)
    margin_stats = stats(margin_values)
    today = datetime.date.today().isoformat()

    lines = [
        f"# 销量表解析 · 数据质量报告",
        f"",
        f"> 生成时间：{today}",
        f"> 输入：`{SRC}`",
        f"> 输出：`{OUT_CSV}`",
        f"",
        f"## 📊 总览",
        f"",
        f"| 指标 | 值 |",
        f"|---|---:|",
        f"| 总 SKU 行数 | **{stat_total}** |",
        f"| 不同品牌数 | {len(stat_by_brand)} |",
        f"| 不同品类数 | {len(stat_by_category)} |",
        f"| 不同店铺数 | {len(stat_by_shop)} |",
        f"| 0 销量 SKU 数（未起量候选）| **{stat_zero_sales}**（{pct(stat_zero_sales, stat_total)}）|",
        f"| 缺品牌字段 | {stat_no_brand} |",
        f"| 缺分类字段 | {stat_no_category} |",
        f"",
        f"## 🏷️ 品牌分布（前 10）",
        f"",
        f"| 品牌 | SKU 数 | 占比 |",
        f"|---|---:|---:|",
    ]
    for brand, cnt in stat_by_brand.most_common(10):
        lines.append(f"| {brand} | {cnt} | {pct(cnt, stat_total)} |")

    lines.extend([
        f"",
        f"## 📦 品类分布（前 15）",
        f"",
        f"| 分类 | SKU 数 | 占比 |",
        f"|---|---:|---:|",
    ])
    for cat, cnt in stat_by_category.most_common(15):
        lines.append(f"| {cat} | {cnt} | {pct(cnt, stat_total)} |")

    lines.extend([
        f"",
        f"## 🏪 店铺分布",
        f"",
        f"| 店铺 | SKU 数 |",
        f"|---|---:|",
    ])
    for shop, cnt in stat_by_shop.most_common():
        lines.append(f"| {shop} | {cnt} |")

    lines.extend([
        f"",
        f"## 💰 实际销售量分布",
        f"",
        f"| 分位数 | 值 |",
        f"|---|---:|",
        f"| 样本数 | {sales_stats['count']} |",
    ])
    if sales_stats["count"] > 0:
        lines.extend([
            f"| 最低 | {sales_stats['min']:.0f} |",
            f"| P25 | {sales_stats['p25']:.0f} |",
            f"| 中位数 | {sales_stats['median']:.0f} |",
            f"| P75 | {sales_stats['p75']:.0f} |",
            f"| 最高 | {sales_stats['max']:.0f} |",
            f"| 均值 | {sales_stats['mean']:.1f} |",
        ])

    lines.extend([
        f"",
        f"## 📈 毛利率%分布",
        f"",
        f"| 分位数 | 值 |",
        f"|---|---:|",
        f"| 样本数 | {margin_stats['count']} |",
    ])
    if margin_stats["count"] > 0:
        lines.extend([
            f"| 最低 | {margin_stats['min']:.2f}% |",
            f"| P25 | {margin_stats['p25']:.2f}% |",
            f"| 中位数 | {margin_stats['median']:.2f}% |",
            f"| P75 | {margin_stats['p75']:.2f}% |",
            f"| 最高 | {margin_stats['max']:.2f}% |",
            f"| 均值 | {margin_stats['mean']:.2f}% |",
        ])

    lines.extend([
        f"",
        f"## ⚠️ 已知问题与下一步",
        f"",
        f"1. 「未起量」阈值待对齐：当前简单按 `实际销售量 = 0` 标记，{stat_zero_sales} 行命中。",
        f"   实际业务可能需要按品类中位数的 30% 作为阈值，或考虑上架时间。",
        f"2. 当前 product_assets.csv 是销量表全量替换，**没有「细分品类(按版型)」语义层**——",
        f"   那一层是货盘独有的。下一步要把货盘的「版型」信息映射到这张表（通过货品编号 JOIN）。",
        f"3. 销量表没有「竞品参考链接」——这部分要从旧版 `data/_legacy/product_assets_v0_pallet.csv`",
        f"   抽出来独立成 `data/competitor_intel.csv`。",
        f"4. 图片字段是 alicdn URL（不是 DISPIMG），可直接预览，不需要图床迁移。",
        f"",
        f"## 🔄 下一步 Pipeline",
        f"",
        f"- Step 2：合并博凯+宏博模块库 → `data/modules.csv`（~700 条）",
        f"- Step 3：展开模块库「复用记录」 → `data/module_product_link.csv`（~1500 条估）",
        f"- Step 4：重写 sediment_pallet.py，只产 `data/competitor_intel.csv`（~300 条估）",
        f"- Step 5：数据质量交叉验证（4 张表 JOIN 通不通）",
    ])

    with open(OUT_REPORT, "w", encoding="utf-8-sig") as f:
        f.write("\n".join(lines))

    print()
    print(f"✅ Step 1 完成")
    print(f"  - {OUT_CSV}  ({stat_total} 行)")
    print(f"  - {OUT_REPORT}")
    print()
    print(f"📊 速览：")
    print(f"  品牌数={len(stat_by_brand)} · 品类数={len(stat_by_category)} · 店铺数={len(stat_by_shop)}")
    print(f"  0 销量 SKU = {stat_zero_sales} ({pct(stat_zero_sales, stat_total)})")
    if sales_values:
        print(f"  销量中位数 = {sales_stats['median']:.0f}，P75 = {sales_stats['p75']:.0f}")


if __name__ == "__main__":
    main()
